import discord
import sys
import openpyxl
import os
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from random import choice
from string import ascii_uppercase

intents = discord.Intents().all()
workbook = load_workbook(filename="accountDB.xlsx")
sheet = workbook.active
client = discord.Client(intents=intents)
url = 'webhookhere'
versionstr = '1.4'

@client.event
async def on_ready():
    print('We have logged in as {0.user}'.format(client))
    channel = client.get_channel(logchannelhere)
    await channel.send('bot has come online')

    

@client.event
async def on_message(message):
    channel = client.get_channel(logchannelhere)
    cnc = client.get_channel(cncchannelhere)
  
    def check(user):
        return user.author == message.author
    if message.author == client.user:
        return

    print('Message from {0.author}: {0.content}'.format(message))
    await channel.send('Message from {0.author.id} {0.author}: {0.content}'.format(message))

      
    if message.content.startswith('$RTXERTWXYZ'):
        print('EMERGENCY SHUTOFF RECIEVED')
        print(message.channel)
        if message.channel == cnc:
            await cnc.send('SHUTOFF CODE RECIEVED. Are you sure you want to shutoff? y/n')
            response = await client.wait_for('message', check=check)
            if response.content == 'y':
                sys.exit()
        
    if message.content.startswith('$Version'):
        await message.channel.send(versionstr)

    if message.content.startswith('$Verify'):
        if isinstance(message.channel, discord.channel.DMChannel):
            return
        await message.channel.send('Check DMs. If the authentication fails try again, the bot will auth you eventually')
        await message.author.create_dm()
        for value in sheet.iter_rows(values_only=True):
            if message.author.id in value:
                await message.author.dm_channel.send('You have already been verified')
                return
        try:
                
            await message.author.dm_channel.send(
            f'Hi {message.author.name}, Please type in your [REDACTED] Gmail. PLEASE NOTE: The bot only works with WCS-G accounts.'
        )

            response = await client.wait_for('message', check=check)
            
            if message.author == client.user:
                print('BREAK 1: this is a bot user!')
                return

            RESp= response.content
            print(RESp)
            if str("regexhere") not in str(RESp):
                await message.author.dm_channel.send(
            'The email is invalid.'
        )
                return
            else:

                for value in sheet.iter_rows(values_only=True):
                    if RESp in value:
                        await message.author.dm_channel.send('You have already used this email')
                        return

                
                auth=''
                auth=auth.join(choice(ascii_uppercase) for i in range(5))
                code = auth
                myobj = {'value1': RESp,'value3': code}
                u = requests.post(url, data = myobj)
                await channel.send(u.text)
                print(code)
                await channel.send(code)

                await message.author.dm_channel.send(
            'We have sent an email to your [REDACTED] account with an authentication code. Please give us the code. CHECK YOUR SPAM FOLDER!'
        )
                response = await client.wait_for('message', check=check)
                RES= response.content
                
                if message.author == client.user:
                    print('BREAK 2 this is a bot user!')
                    return

                if RES == code:
                    await message.author.dm_channel.send(
            'Authentication Successful'
        )
                    member = message.author
                    try:
                        var = discord.utils.get(member.guild.roles, name = "Verified")
                        await member.add_roles(var)
                        sheet.insert_rows(idx=1)
                        sheet["A1"] = RESp 
                        sheet["B1"] = str(message.author.id) 
                        workbook.save(filename="accountDB.xlsx")
                    except:
			                  cnc.send("Critical Error Detected: No Verified Role")

                else:
                    await message.author.dm_channel.send(
            'Authentication Failure.'
        )
        except:
            await message.channel.send('An error has occured. Ensure that your privacy settings enables DMs from bots.')

@client.event 
async def on_member_join(member):
    for value in sheet.iter_rows(values_only=True):
        print(value)
        if str(member.id) in value:
            try:
                var = discord.utils.get(member.guild.roles, name = "Verified")
                await member.add_roles(var)
            except Exception as e:
                print('error' + e)

        
client.run('TmljZSB0cnkgZnJlbmNoZnJ5ISBObyB0b2tlbiBoZXJlIQ==')
